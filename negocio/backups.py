import zipfile

import openpyxl
from PyQt6 import QtSql
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from bbdd.modelos import Vehiculo, Cliente
from negocio import dni


class ServicioBackup:
	@staticmethod
	def restaurar_copia(ruta: str) -> bool:
		zipf = zipfile.ZipFile(ruta, 'r')
		zipf.extractall()
		zipf.close()
		return True

	@staticmethod
	def hacer_copia(ruta: str) -> bool:
		zipf = zipfile.ZipFile(ruta, 'w', zipfile.ZIP_DEFLATED)
		zipf.write("bbdd.sqlite")
		zipf.close()
		return True

	def exportar_excel(self, ruta: str, clientes: bool, coches: bool, servicios: bool,
					   incluir_historico: bool = False) -> bool:
		try:
			wb = openpyxl.Workbook()
			wb.remove_sheet(wb.active)
			if clientes:
				self.exportar_clientes_excel(wb, incluir_historico)
			if coches:
				self.exportar_coches_excel(wb, incluir_historico)
			if servicios:
				self.exportar_servicios_excel(wb, incluir_historico)
			wb.save(ruta)
		except Exception as error:
			print("Error al exportar a excel: ", error)
			return False

	@staticmethod
	def exportar_clientes_excel(wb: openpyxl.Workbook, historico: bool) -> bool:
		try:
			hoja_clientes: Worksheet = wb.create_sheet("clientes")
			elementos = ["DNI", "Nombre", "Fecha alta", "Direccion", "Provincia", "Municipio",
						 "Admite efectivo", "Admite factura", "Admite transferencia",
						 "Fecha de baja"]
			hoja_clientes.append(elementos)

			query = QtSql.QSqlQuery()
			if historico:
				query.exec("SELECT * FROM clientes")
			else:
				query.prepare("SELECT * FROM clientes WHERE fecha_baja IS NULL ORDER BY alta")

			if query.exec():
				fila = 2
				while query.next():
					for i in range(1, 11):
						hoja_clientes.cell(column=i, row=fila, value=query.value(i - 1))
					fila += 1

			return True
		except Exception as error:
			print("Error al exportar a excel: ", error)
			return False

	@staticmethod
	def exportar_coches_excel(wb: openpyxl.Workbook, historico: bool) -> bool:
		try:
			hoja_coches: Worksheet = wb.create_sheet("coches")

			hoja_coches.append(["Matricula", "DNI", "Marca", "Modelo", "Motor", "Fecha de baja"])

			query = QtSql.QSqlQuery()
			if historico:
				query.exec("SELECT * FROM coches")
			else:
				query.prepare("SELECT * FROM coches WHERE fecha_baja IS NULL ORDER BY matricula")

			if query.exec():
				fila = 2
				while query.next():
					for i in range(1, 7):
						hoja_coches.cell(column=i, row=fila, value=query.value(i - 1))
					fila += 1
			return True
		except Exception as error:
			print("Error al exportar a excel: ", error)
			return False

	@staticmethod
	def exportar_servicios_excel(wb: openpyxl.Workbook, historico: bool) -> bool:
		try:
			hoja_servicios: Worksheet = wb.create_sheet("servicios")

			hoja_servicios.append(["ID", "Nombre", "Precio unitario", "Fecha de alta", "Fecha de modificación", "Fecha de baja"])

			query = QtSql.QSqlQuery()
			if historico:
				query.exec("SELECT * FROM servicios")
			else:
				query.prepare("SELECT * FROM servicios WHERE fechaBaja IS NULL")

			if query.exec():
				fila = 2
				while query.next():
					for i in range(1, 7):
						hoja_servicios.cell(column=i, row=fila, value=query.value(i - 1))
					fila += 1
			return True
		except Exception as error:
			print("Error al exportar a excel: ", error)
			return False

	@staticmethod
	def comprobar_tipos_importables_excel(ruta) -> [bool, bool]:
		try:
			book: Workbook = openpyxl.load_workbook(ruta)
			hojas = book.get_sheet_names()

			# resultado = NoClientes y NoCoches
			resultado = (False, False)

			for hoja in hojas:
				if hoja == "clientes":
					# resultado = SiClientes y LoqueseaCoches
					resultado = (True, resultado[1])
				if hoja == "coches":
					# resultado = LoqueseaClientes y SiCoches
					resultado = (resultado[0], True)
			return resultado
		except Exception as error:
			print("Error al comprobar tipos importables: ", error)
			return False, False

	def importar_excel(self, ruta: str, clientes: bool, coches: bool) -> bool:
		try:
			book: Workbook = openpyxl.load_workbook(ruta)
			if clientes:
				self.importar_clientes_excel(book["clientes"])
			if coches:
				self.importar_coches_excel(book["coches"])
		except Exception as error:
			print("Error al exportar a excel: ", error)
			return False

	@staticmethod
	def importar_clientes_excel(sheet: Worksheet) -> bool:
		try:
			filas: list[Cliente] = []
			for fila in sheet.iter_rows(min_row=2):
				if not dni.validar(fila[0].value):
					print("DNI no valido: ", fila[0].value)
					continue

				filas.append(Cliente(
					fila[0].value,
					fila[1].value,
					fila[2].value,
					fila[3].value,
					fila[4].value,
					fila[5].value,
					fila[6].value,
					fila[7].value,
					fila[8].value
				))

			query = QtSql.QSqlQuery()

			for fila in filas:
				query.prepare("INSERT INTO clientes VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, NULL)")
				query.addBindValue(fila.dni)
				query.addBindValue(fila.nombre)
				query.addBindValue(fila.fecha_alta.strftime("%Y-%m-%d"))
				query.addBindValue(fila.direccion)
				query.addBindValue(fila.provincia)
				query.addBindValue(fila.municipio)
				query.addBindValue(fila.efectivo)
				query.addBindValue(fila.factura)
				query.addBindValue(fila.transferencia)
				if not (query.exec()):
					print("Error al importar cliente: " + query.lastError().text())
			return True

		except Exception as error:
			print("Error al importar clientes excel: ", error)
			return False

	@staticmethod
	def importar_coches_excel(sheet: Worksheet) -> bool:
		try:
			filas: list[Vehiculo] = []
			for fila in sheet.iter_rows(min_row=2):  # Empieza en 1 para saltarse la cabecera

				filas.append(Vehiculo(
					fila[0].value,
					fila[1].value,
					fila[2].value,
					fila[3].value,
					fila[4].value
				))

			query = QtSql.QSqlQuery()
			query.prepare(
				"INSERT INTO coches (matricula, dnicli, marca, modelo, motor) VALUES (?, ?, ?, ?, ?)")

			for fila in filas:
				query.addBindValue(fila.matricula)
				query.addBindValue(fila.dni)
				query.addBindValue(fila.marca)
				query.addBindValue(fila.modelo)
				query.addBindValue(fila.motor)
				query.exec()

			return True
		except Exception as error:
			print("Error al importar de excel: ", error)
			return False
